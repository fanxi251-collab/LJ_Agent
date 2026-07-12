from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = "tourism_analytics_v1"
REQUIRED_HEADERS = (
    "tourist_id",
    "user_nickname",
    "age",
    "gender",
    "attraction_name",
    "attraction_content",
    "attraction_type",
    "visit_date",
    "stay_duration",
    "ticket_cost",
    "food_cost",
    "shopping_cost",
    "transport_cost",
    "entertainment_cost",
    "total_cost",
    "group_size",
    "satisfaction",
)
NUMERIC_FIELDS = (
    "age",
    "stay_duration",
    "ticket_cost",
    "food_cost",
    "shopping_cost",
    "transport_cost",
    "entertainment_cost",
    "total_cost",
    "group_size",
    "satisfaction",
)
COST_FIELDS = (
    "ticket_cost",
    "food_cost",
    "shopping_cost",
    "transport_cost",
    "entertainment_cost",
)
PRIVATE_FIELDS = {"tourist_id", "user_nickname", "attraction_content"}
MONEY_QUANTUM = Decimal("0.01")


class AnalyticsDataError(ValueError):
    """Reject unreliable source data so the dashboard never presents partial statistics."""


def build_snapshot(
    source_path: str | Path,
    output_path: str | Path | None = None,
    *,
    min_rank_visits: int = 100,
) -> dict[str, Any]:
    source = Path(source_path)
    if not source.is_file():
        raise AnalyticsDataError(f"找不到 Excel 文件：{source}")
    if min_rank_visits < 1:
        raise AnalyticsDataError("排行榜样本门槛必须大于 0。")

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise AnalyticsDataError(f"无法读取 Excel 文件：{exc}") from exc

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise AnalyticsDataError("Excel 文件没有表头。")
        header_map = _header_map(headers)
        records = [_parse_row(row, excel_row, header_map) for excel_row, row in enumerate(rows, 2)]
    finally:
        workbook.close()

    if not records:
        raise AnalyticsDataError("Excel 文件没有可分析的数据行。")

    snapshot = _aggregate_snapshot(source, records, min_rank_visits)
    if output_path is not None:
        _write_snapshot_atomically(Path(output_path), snapshot)
    return snapshot


def _header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    normalized = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    missing = [field for field in REQUIRED_HEADERS if field not in normalized]
    if missing:
        raise AnalyticsDataError(f"Excel 缺少必需字段：{', '.join(missing)}")
    return {field: normalized[field] for field in REQUIRED_HEADERS}


def _parse_row(row: tuple[Any, ...], excel_row: int, header_map: dict[str, int]) -> dict[str, Any]:
    values = {
        field: row[index] if index < len(row) else None
        for field, index in header_map.items()
    }
    required_text = ("tourist_id", "gender", "attraction_name", "attraction_type")
    for field in required_text:
        if values[field] is None or not str(values[field]).strip():
            raise AnalyticsDataError(f"第 {excel_row} 行字段 {field} 不能为空。")

    parsed: dict[str, Any] = {
        "tourist_id": str(values["tourist_id"]).strip(),
        "gender": str(values["gender"]).strip(),
        "attraction_name": str(values["attraction_name"]).strip(),
        "attraction_type": str(values["attraction_type"]).strip(),
        "visit_date": _parse_date(values["visit_date"], excel_row),
    }
    for field in NUMERIC_FIELDS:
        parsed[field] = _parse_decimal(values[field], excel_row, field)

    age = int(parsed["age"])
    group_size = int(parsed["group_size"])
    satisfaction = int(parsed["satisfaction"])
    if age < 0 or group_size < 1 or satisfaction not in {1, 2, 3, 4, 5}:
        raise AnalyticsDataError(f"第 {excel_row} 行年龄、同行人数或满意度超出允许范围。")
    parsed["age"] = age
    parsed["group_size"] = group_size
    parsed["satisfaction"] = satisfaction
    parsed["missing_cells"] = sum(values[field] in (None, "") for field in REQUIRED_HEADERS)
    return parsed


def _parse_date(value: Any, excel_row: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value).strip()).date()
    except (TypeError, ValueError) as exc:
        raise AnalyticsDataError(f"第 {excel_row} 行字段 visit_date 不是有效日期。") from exc


def _parse_decimal(value: Any, excel_row: int, field: str) -> Decimal:
    try:
        if value is None or str(value).strip() == "":
            raise InvalidOperation
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise AnalyticsDataError(f"第 {excel_row} 行字段 {field} 不是有效数值。") from exc


def _aggregate_snapshot(
    source: Path,
    records: list[dict[str, Any]],
    min_rank_visits: int,
) -> dict[str, Any]:
    tourist_visits = Counter(record["tourist_id"] for record in records)
    attraction_stats: dict[str, dict[str, Any]] = defaultdict(_new_group_stats)
    type_stats: dict[str, dict[str, Any]] = defaultdict(_new_group_stats)
    months: Counter[str] = Counter()
    ages: Counter[str] = Counter()
    genders: Counter[str] = Counter()
    group_sizes: Counter[str] = Counter()
    satisfaction_counts: Counter[int] = Counter()
    consumption_totals = {field: Decimal("0") for field in COST_FIELDS}
    analytical_rows: Counter[tuple[Any, ...]] = Counter()
    total_stay = Decimal("0")
    total_cost = Decimal("0")
    total_satisfaction = Decimal("0")
    mismatch_rows = 0
    missing_cells = 0

    for record in records:
        months[record["visit_date"].strftime("%Y-%m")] += 1
        ages[_age_group(record["age"])] += 1
        genders[record["gender"]] += 1
        group_sizes[_group_size_label(record["group_size"])] += 1
        satisfaction_counts[record["satisfaction"]] += 1
        total_stay += record["stay_duration"]
        total_cost += record["total_cost"]
        total_satisfaction += Decimal(record["satisfaction"])
        missing_cells += record["missing_cells"]
        component_total = sum((record[field] for field in COST_FIELDS), Decimal("0"))
        if abs(_money(component_total) - _money(record["total_cost"])) > MONEY_QUANTUM:
            mismatch_rows += 1
        for field in COST_FIELDS:
            consumption_totals[field] += record[field]
        _add_group_record(attraction_stats[record["attraction_name"]], record)
        _add_group_record(type_stats[record["attraction_type"]], record)
        analytical_rows[_analytical_key(record)] += 1

    count = len(records)
    attraction_rows = _group_rows(attraction_stats)
    type_rows = _group_rows(type_stats)
    repeat_rate = sum(visits > 1 for visits in tourist_visits.values()) / len(tourist_visits)
    attraction_rankings = _build_attraction_rankings(attraction_rows, min_rank_visits)
    average_cost = float(total_cost / count)
    average_satisfaction = float(total_satisfaction / count)
    insights = _build_insights(months, ages, type_rows, average_cost, average_satisfaction)

    return {
        "metadata": {
            "schema_version": SCHEMA_VERSION,
            "source_filename": source.name,
            "source_sha256": _sha256(source),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "period_start": min(record["visit_date"] for record in records).isoformat(),
            "period_end": max(record["visit_date"] for record in records).isoformat(),
            "row_count": count,
            "static_demo_data": True,
        },
        "kpis": {
            "visit_count": count,
            "tourist_count": len(tourist_visits),
            "attraction_count": len(attraction_stats),
            "average_stay_hours": _rounded(total_stay / count),
            "average_total_cost": _rounded(total_cost / count),
            "average_satisfaction": _rounded(total_satisfaction / count),
        },
        "monthly_trend": [
            {"month": month, "visit_count": months[month]} for month in sorted(months)
        ],
        "demographics": {
            "age_groups": [
                {"label": label, "count": ages[label]}
                for label in ("≤18", "19-25", "26-35", "36-45", "46-60", "60+")
            ],
            "genders": [
                {"label": label, "count": value} for label, value in _sorted_counter(genders)
            ],
            "group_sizes": [
                {"label": label, "count": group_sizes[label]}
                for label in ("1人", "2人", "3人", "4人", "5人", "6人及以上")
            ],
            "repeat_visitor_rate": round(repeat_rate, 4),
        },
        "attraction_types": sorted(type_rows, key=lambda item: (-item["visit_count"], item["name"])),
        "attraction_rankings": attraction_rankings,
        "consumption": {
            "categories": [
                {"key": field, "label": _cost_label(field), "total": _rounded(consumption_totals[field])}
                for field in COST_FIELDS
            ],
            "by_attraction_type": sorted(
                [
                    {
                        "name": item["name"],
                        "visit_count": item["visit_count"],
                        "average_total_cost": item["average_total_cost"],
                    }
                    for item in type_rows
                ],
                key=lambda item: (-item["average_total_cost"], item["name"]),
            ),
        },
        "satisfaction": {
            "distribution": [
                {"score": score, "count": satisfaction_counts[score]} for score in range(1, 6)
            ],
            "quadrants": [
                {
                    "name": item["name"],
                    "visit_count": item["visit_count"],
                    "average_total_cost": item["average_total_cost"],
                    "average_satisfaction": item["average_satisfaction"],
                }
                for item in type_rows
            ],
            "overall_average_cost": _rounded(Decimal(str(average_cost))),
            "overall_average_satisfaction": _rounded(Decimal(str(average_satisfaction))),
        },
        "insights": insights,
        "quality": {
            "missing_cells": missing_cells,
            "analytical_duplicate_rows": sum(count - 1 for count in analytical_rows.values() if count > 1),
            "invalid_rows": 0,
            "total_cost_mismatch_rows": mismatch_rows,
            "rank_minimum_visits": min_rank_visits,
        },
    }


def _new_group_stats() -> dict[str, Any]:
    return {
        "visit_count": 0,
        "tourist_ids": set(),
        "total_stay": Decimal("0"),
        "total_cost": Decimal("0"),
        "total_satisfaction": Decimal("0"),
        "attraction_type": "",
    }


def _add_group_record(stats: dict[str, Any], record: dict[str, Any]) -> None:
    stats["visit_count"] += 1
    stats["tourist_ids"].add(record["tourist_id"])
    stats["total_stay"] += record["stay_duration"]
    stats["total_cost"] += record["total_cost"]
    stats["total_satisfaction"] += Decimal(record["satisfaction"])
    stats["attraction_type"] = record["attraction_type"]


def _group_rows(groups: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for name, stats in groups.items():
        count = stats["visit_count"]
        result.append(
            {
                "name": name,
                "attraction_type": stats["attraction_type"],
                "visit_count": count,
                "tourist_count": len(stats["tourist_ids"]),
                "average_stay_hours": _rounded(stats["total_stay"] / count),
                "average_total_cost": _rounded(stats["total_cost"] / count),
                "average_satisfaction": _rounded(stats["total_satisfaction"] / count),
            }
        )
    return result


def _build_attraction_rankings(
    attraction_rows: list[dict[str, Any]],
    min_rank_visits: int,
) -> dict[str, list[dict[str, Any]]]:
    eligible = [item for item in attraction_rows if item["visit_count"] >= min_rank_visits]
    return {
        "popular": sorted(
            attraction_rows, key=lambda item: (-item["visit_count"], item["name"])
        )[:10],
        "high_spend": sorted(
            eligible,
            key=lambda item: (-item["average_total_cost"], -item["visit_count"], item["name"]),
        )[:10],
        "long_stay": sorted(
            eligible,
            key=lambda item: (-item["average_stay_hours"], -item["visit_count"], item["name"]),
        )[:10],
        "high_satisfaction": sorted(
            eligible,
            key=lambda item: (-item["average_satisfaction"], -item["visit_count"], item["name"]),
        )[:10],
        "low_satisfaction": sorted(
            eligible,
            key=lambda item: (item["average_satisfaction"], -item["visit_count"], item["name"]),
        )[:10],
    }


def _build_insights(
    months: Counter[str],
    ages: Counter[str],
    type_rows: list[dict[str, Any]],
    overall_cost: float,
    overall_satisfaction: float,
) -> list[dict[str, str]]:
    peak_month, peak_count = sorted(months.items(), key=lambda item: (-item[1], item[0]))[0]
    age_group, age_count = sorted(ages.items(), key=lambda item: (-item[1], item[0]))[0]
    high_cost_low_score = sorted(
        [
            item
            for item in type_rows
            if item["average_total_cost"] > overall_cost
            and item["average_satisfaction"] < overall_satisfaction
        ],
        key=lambda item: (-item["average_total_cost"], item["name"]),
    )
    low_cost_high_score = sorted(
        [
            item
            for item in type_rows
            if item["average_total_cost"] < overall_cost
            and item["average_satisfaction"] > overall_satisfaction
        ],
        key=lambda item: (-item["average_satisfaction"], item["name"]),
    )
    insights = [
        {
            "kind": "trend",
            "title": "年度客流峰值",
            "description": f"{peak_month} 的游览记录最多，共 {peak_count} 条。",
        },
        {
            "kind": "audience",
            "title": "核心年龄客群",
            "description": f"{age_group} 岁年龄段记录最多，共 {age_count} 条。",
        },
    ]
    if high_cost_low_score:
        item = high_cost_low_score[0]
        insights.append(
            {
                "kind": "warning",
                "title": "高消费低满意类型",
                "description": f"{item['name']}平均消费较高但满意度低于整体，建议优先检查价格感知和服务体验。",
            }
        )
    if low_cost_high_score:
        item = low_cost_high_score[0]
        insights.append(
            {
                "kind": "opportunity",
                "title": "高口碑引流类型",
                "description": f"{item['name']}满意度高且平均消费较低，可作为游客端重点推荐内容。",
            }
        )
    return insights


def _analytical_key(record: dict[str, Any]) -> tuple[Any, ...]:
    return tuple(
        record[field]
        for field in REQUIRED_HEADERS
        if field not in PRIVATE_FIELDS and field in record
    )


def _age_group(age: int) -> str:
    if age <= 18:
        return "≤18"
    if age <= 25:
        return "19-25"
    if age <= 35:
        return "26-35"
    if age <= 45:
        return "36-45"
    if age <= 60:
        return "46-60"
    return "60+"


def _group_size_label(group_size: int) -> str:
    return f"{group_size}人" if group_size <= 5 else "6人及以上"


def _cost_label(field: str) -> str:
    return {
        "ticket_cost": "门票",
        "food_cost": "餐饮",
        "shopping_cost": "购物",
        "transport_cost": "交通",
        "entertainment_cost": "娱乐",
    }[field]


def _sorted_counter(counter: Counter[str]) -> list[tuple[str, int]]:
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))


def _rounded(value: Decimal) -> float:
    return float(value.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP))


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_snapshot_atomically(output: Path, snapshot: dict[str, Any]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(output.suffix + ".tmp")
    temporary.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="从旅游行为 Excel 生成管理端固定分析快照。")
    parser.add_argument("--input", required=True, help="源 Excel 文件路径")
    parser.add_argument(
        "--output",
        default="data/tourism_analytics_snapshot.json",
        help="快照输出路径",
    )
    parser.add_argument("--min-rank-visits", type=int, default=100, help="排行榜最小样本数")
    args = parser.parse_args()
    try:
        snapshot = build_snapshot(args.input, args.output, min_rank_visits=args.min_rank_visits)
    except AnalyticsDataError as exc:
        parser.error(str(exc))
    metadata = snapshot["metadata"]
    print(
        f"快照已生成：{args.output}；记录 {metadata['row_count']} 条；"
        f"周期 {metadata['period_start']} 至 {metadata['period_end']}。"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
